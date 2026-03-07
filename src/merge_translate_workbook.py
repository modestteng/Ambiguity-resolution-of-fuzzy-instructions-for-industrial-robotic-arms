from __future__ import annotations

import argparse
import json
import re
import time
from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path
from typing import Iterable

import requests
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


ASCII_RE = re.compile(r"[A-Za-z]")
WHITESPACE_RE = re.compile(r"\s+")

SHEET_NAME_MAP = {
    "paper_matrix": "论文主表",
    "batch_review": "分批回顾",
    "method_tree": "方法分类树",
    "year_trend": "年份趋势",
    "gap_analysis": "空白区分析",
    "candidate_topics": "候选选题",
    "summary": "汇总统计",
}


def clean_text(text: str) -> str:
    return WHITESPACE_RE.sub(" ", text).strip()


def needs_translation(value: object) -> bool:
    return isinstance(value, str) and bool(ASCII_RE.search(value))


def cache_path(root: Path) -> Path:
    data_dir = root / "data"
    data_dir.mkdir(exist_ok=True)
    return data_dir / "translation_cache_zh.json"


def load_cache(path: Path) -> dict[str, str]:
    if not path.exists():
        return {}
    try:
        return json.loads(path.read_text(encoding="utf-8"))
    except Exception:
        return {}


def save_cache(path: Path, cache: dict[str, str]) -> None:
    path.write_text(json.dumps(cache, ensure_ascii=False, indent=2), encoding="utf-8")


def translate_text(session: requests.Session, text: str, retries: int = 4) -> str:
    cleaned = clean_text(text)
    if not cleaned:
        return cleaned
    params = {
        "client": "gtx",
        "sl": "auto",
        "tl": "zh-CN",
        "dt": "t",
        "q": cleaned,
    }
    for attempt in range(retries):
        try:
            response = session.get(
                "https://translate.googleapis.com/translate_a/single",
                params=params,
                timeout=30,
            )
            response.raise_for_status()
            data = response.json()
            segments = data[0] if data and data[0] else []
            translated = "".join(part[0] for part in segments if part and part[0])
            return translated or cleaned
        except Exception:
            time.sleep(1.5 * (attempt + 1))
    return cleaned


def collect_unique_strings(workbook_path: Path) -> list[str]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    seen: set[str] = set()
    ordered: list[str] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        if ASCII_RE.search(sheet_name):
            label = SHEET_NAME_MAP.get(sheet_name, sheet_name)
            if label != sheet_name:
                seen.add(sheet_name)
        for row in ws.iter_rows(values_only=True):
            for value in row:
                if needs_translation(value):
                    text = clean_text(str(value))
                    if text and text not in seen:
                        seen.add(text)
                        ordered.append(text)
    return ordered


def translate_all(strings: Iterable[str], cache: dict[str, str], workers: int = 6) -> dict[str, str]:
    pending = [text for text in strings if text not in cache]
    if not pending:
        return cache

    def run_one(text: str) -> tuple[str, str]:
        with requests.Session() as session:
            session.headers.update({"User-Agent": "Mozilla/5.0"})
            return text, translate_text(session, text)

    with ThreadPoolExecutor(max_workers=workers) as executor:
        futures = {executor.submit(run_one, text): text for text in pending}
        done = 0
        total = len(pending)
        for future in as_completed(futures):
            src, translated = future.result()
            cache[src] = translated
            done += 1
            if done % 100 == 0 or done == total:
                print(f"Translated {done}/{total} strings")
    return cache


def translated_value(value: object, cache: dict[str, str]) -> object:
    if not needs_translation(value):
        return value
    text = clean_text(str(value))
    return cache.get(text, text)


def style_sheet(ws) -> None:
    header_fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    section_fill = PatternFill(fill_type="solid", fgColor="FDE9D9")
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = header_fill
    for row in ws.iter_rows():
        if row[0].value == "区块":
            for cell in row:
                cell.font = Font(bold=True)
                cell.fill = section_fill
    ws.freeze_panes = "A2"
    for idx, column in enumerate(ws.columns, start=1):
        values = [str(cell.value) if cell.value is not None else "" for cell in column]
        width = min(max((len(v) for v in values), default=12) + 2, 55)
        width = max(width, 12)
        ws.column_dimensions[get_column_letter(idx)].width = width


def build_merged_sheet(source_path: Path, output_path: Path, cache: dict[str, str]) -> None:
    src_wb = load_workbook(source_path, read_only=True, data_only=True)
    out_wb = Workbook()
    ws = out_wb.active
    ws.title = "大表中文"
    ws.append(["区块", "列1", "列2", "列3", "列4", "列5", "列6", "列7", "列8", "列9", "列10", "列11", "列12", "列13", "列14", "列15", "列16", "列17", "列18", "列19", "列20", "列21", "列22", "列23", "列24"])

    for sheet_name in src_wb.sheetnames:
        translated_sheet = SHEET_NAME_MAP.get(sheet_name, cache.get(sheet_name, sheet_name))
        ws.append(["区块", translated_sheet])
        src_ws = src_wb[sheet_name]
        for row in src_ws.iter_rows(values_only=True):
            translated_row = [translated_sheet]
            translated_row.extend(translated_value(value, cache) for value in row)
            ws.append(translated_row)
        ws.append([])

    style_sheet(ws)
    out_wb.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Merge workbook sheets vertically and translate English text into Chinese.")
    parser.add_argument("--root", type=Path, default=Path.cwd(), help="Project root.")
    parser.add_argument(
        "--input",
        type=Path,
        default=Path.cwd().parent / "具身模糊指令文献矩阵_总表.xlsx",
        help="Source xlsx path.",
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=Path.cwd().parent / "具身模糊指令文献矩阵_单表中文.xlsx",
        help="Merged Chinese xlsx path.",
    )
    args = parser.parse_args()

    cpath = cache_path(args.root)
    cache = load_cache(cpath)
    unique_strings = collect_unique_strings(args.input)
    print(f"Unique strings needing translation: {len(unique_strings)}")
    cache = translate_all(unique_strings, cache)
    save_cache(cpath, cache)
    build_merged_sheet(args.input, args.output, cache)
    print(f"Merged workbook saved to: {args.output}")


if __name__ == "__main__":
    main()
