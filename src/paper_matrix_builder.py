from __future__ import annotations

import argparse
import math
import re
from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Iterable

from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from pypdf import PdfReader


TITLE_SPLIT_RE = re.compile(
    r"\s+-\s+(?:arXiv|ArXiv\.org|openalex|Cornell University|IEEE|Findings|Lecture notes|Journal|CAAI|Biological|Complex|Advanced|Proceedings|International|Neural|Human).*",
    re.IGNORECASE,
)
ABSTRACT_START_RE = re.compile(r"\babstract\b[:\s-]*", re.IGNORECASE)
ABSTRACT_END_RE = re.compile(
    r"\b(?:1\.?\s*introduction|i+\.\s*introduction|keywords|index terms|introduction)\b",
    re.IGNORECASE,
)
SECTION_RE = re.compile(
    r"(?im)^\s*(?:\d+(?:\.\d+)*\.?\s+|[IVXLC]+\.\s+)?"
    r"(conclusion(?:s)?(?:\s+and\s+future\s+work)?|discussion|limitations?|future work)\s*$"
)
SENTENCE_SPLIT_RE = re.compile(r"(?<=[.!?])\s+")
UPPER_TOKEN_RE = re.compile(r"^[A-Z0-9][A-Z0-9\-]{1,20}$")
ILLEGAL_EXCEL_RE = re.compile(r"[\x00-\x08\x0B-\x0C\x0E-\x1F]")


METRIC_TERMS = [
    "accuracy",
    "f1",
    "precision",
    "recall",
    "success rate",
    "success",
    "iou",
    "miou",
    "top-1",
    "top-5",
    "bleu",
    "rouge",
    "cidEr".lower(),
]

DATASET_TERMS = [
    "dataset",
    "datasets",
    "benchmark",
    "benchmarks",
    "evaluated on",
    "evaluate on",
    "validated on",
    "simulation",
    "tabletop",
    "real-world",
    "environment",
    "environments",
]


DISAMBIGUATION_METHOD_RULES = [
    ("对话式澄清", ["clarif", "question", "dialog", "conversation", "multi-turn"]),
    ("上下文推理", ["context", "situational", "history", "memory", "preference", "intent"]),
    ("多模态融合", ["multimodal", "vision-language", "audio-visual", "crossmodal", "fusion"]),
    ("不确定度驱动", ["uncertainty", "confidence", "risk", "abstain", "calibrat"]),
    ("场景图/空间推理", ["scene graph", "3d", "spatial", "position", "graph"]),
    ("反馈式修正", ["feedback", "replan", "error recovery", "correct", "repair"]),
    ("微调/参数高效适配", ["lora", "fine-tun", "parameter-efficient", "peft", "adapter", "prompt tuning"]),
    ("数据集/基准", ["benchmark", "dataset", "survey", "roadmap"]),
]

ARCHITECTURE_RULES = [
    ("LLM", ["large language model", "llm", "gpt", "qwen"]),
    ("VLM/MLLM", ["vision-language model", "vlm", "mllm", "multimodal llm", "clip"]),
    ("VLA/机器人策略", ["vision-language-action", "vla", "policy", "manipulation policy", "grasp policy"]),
    ("神经符号/场景图", ["scene graph", "neuro-symbolic", "symbolic", "graph"]),
    ("扩散/生成式策略", ["diffusion", "generative model"]),
    ("RL/POMDP", ["reinforcement learning", "pomdp", "mdp"]),
    ("分类器/排序器", ["classifier", "classification", "rank", "ranking"]),
    ("综述/数据资源", ["survey", "dataset", "benchmark", "roadmap"]),
]

SCENE_RULES = [
    ("工业机械臂/装配", ["industrial", "assembly", "factory", "manufacturing", "robot arm"]),
    ("桌面抓取/搬运", ["tabletop", "pick-and-place", "grasp", "fetch-and-carry", "clutter"]),
    ("家庭/服务机器人", ["household", "domestic", "service robot", "kitchen", "home"]),
    ("导航/VLN", ["navigation", "nav", "wayfinding", "vln"]),
    ("手术/医疗协作", ["surgical", "surgery", "healthcare", "clinical"]),
    ("多机器人/团队", ["multi-robot", "robot teams", "distributed"]),
    ("无人机/空中平台", ["uav", "aerial", "drone"]),
    ("GUI/数字代理", ["computer-use", "mobile agent"]),
]

AMBIGUITY_RULES = [
    ("指代歧义", ["refer", "referred object", "referring", "which object", "target object"]),
    ("空间歧义", ["spatial", "left", "right", "position", "3d", "location"]),
    ("动作/参数歧义", ["trajectory", "motion", "parameter", "move", "control", "pose", "orientation"]),
    ("目标/任务歧义", ["task ambiguity", "goal", "underspecified", "high-level command"]),
    ("偏好/意图歧义", ["preference", "intent", "desire", "habit"]),
    ("可行性/约束歧义", ["infeasible", "affordance", "constraint", "feasible", "safety"]),
    ("泛化模糊指令", ["ambigu", "vague", "fuzzy", "imperfect language"]),
]

MODALITY_RULES = [
    ("文本", ["language only", "text-only"]),
    ("图像+文本", ["image", "vision", "rgb", "visual", "multimodal"]),
    ("视频/时序视觉+文本", ["video", "egocentric", "temporal", "trajectory"]),
    ("音频+视觉+文本", ["audio", "speech", "prosody", "audio-visual"]),
    ("3D/场景状态+文本", ["3d", "scene graph", "scene description", "state feedback", "point cloud"]),
    ("对话历史+场景", ["dialog", "conversation", "history", "multi-turn"]),
]

HIGH_RELEVANCE_TERMS = [
    "ambigu",
    "disambigu",
    "clarif",
    "fuzzy",
    "instruction",
    "robot",
    "manipulation",
    "grasp",
    "grounding",
    "industrial",
    "embodied",
]

GAP_CANDIDATES = [
    ("工业机械臂/装配", "微调/参数高效适配"),
    ("工业机械臂/装配", "多模态融合"),
    ("工业机械臂/装配", "对话式澄清"),
    ("桌面抓取/搬运", "微调/参数高效适配"),
    ("桌面抓取/搬运", "不确定度驱动"),
    ("家庭/服务机器人", "对话式澄清"),
    ("导航/VLN", "不确定度驱动"),
    ("手术/医疗协作", "可行性/约束歧义"),
]


@dataclass
class PaperRecord:
    index: int
    batch_id: int
    filename: str
    full_title: str
    short_title: str
    year: str
    task_scene: str
    ambiguity_type: str
    disambiguation_method: str
    modality_input: str
    model_architecture: str
    dataset_summary: str
    core_metrics: str
    limitation_summary: str
    inspiration: str
    theme: str
    relevance: str
    research_task: str
    abstract_excerpt: str
    method_evidence: str
    limitation_evidence: str
    evidence_pages: str
    parse_status: str
    notes: str


def clean_text(text: str) -> str:
    text = text.replace("\x00", " ")
    text = text.replace("-\n", "")
    text = re.sub(r"\s+", " ", text)
    text = ILLEGAL_EXCEL_RE.sub("", text)
    return text.strip()


def trim_text(text: str, max_len: int = 260) -> str:
    text = clean_text(text)
    if len(text) <= max_len:
        return text
    return text[: max_len - 3].rstrip() + "..."


def split_sentences(text: str) -> list[str]:
    text = clean_text(text)
    if not text:
        return []
    return [s.strip() for s in SENTENCE_SPLIT_RE.split(text) if len(s.strip()) > 20]


def pick_sentences(text: str, patterns: Iterable[str], limit: int = 2) -> list[str]:
    chosen = []
    for sentence in split_sentences(text):
        low = sentence.lower()
        if any(pattern in low for pattern in patterns):
            chosen.append(sentence)
        if len(chosen) >= limit:
            break
    return chosen


def score_rule(text: str, rules: list[tuple[str, list[str]]], default: str, max_labels: int = 2) -> str:
    low = text.lower()
    scores: list[tuple[int, str]] = []
    for label, keywords in rules:
        score = sum(1 for kw in keywords if kw in low)
        if score > 0:
            scores.append((score, label))
    if not scores:
        return default
    scores.sort(key=lambda item: (-item[0], item[1]))
    top_score = scores[0][0]
    top_labels = [label for score, label in scores if score == top_score][:max_labels]
    return " + ".join(top_labels)


def derive_title_from_filename(filename: str) -> str:
    stem = Path(filename).stem
    title = TITLE_SPLIT_RE.sub("", stem).strip()
    return title or stem


def derive_short_title(full_title: str) -> str:
    prefix = full_title.split(":")[0].strip()
    if 2 <= len(prefix) <= 28:
        return prefix
    tokens = re.findall(r"[A-Za-z0-9\-]+", full_title)
    if tokens and UPPER_TOKEN_RE.match(tokens[0]):
        return tokens[0]
    return " ".join(tokens[:5])[:40] or full_title[:40]


def year_from_filename(filename: str) -> str:
    years = re.findall(r"(20\d{2})", filename)
    return years[-1] if years else ""


def find_pdf_folder(root: Path) -> Path:
    for path in root.iterdir():
        if path.is_dir() and "PDF" in path.name:
            return path
    raise FileNotFoundError("PDF folder not found.")


def windows_safe_path(path: Path) -> str:
    resolved = str(path.resolve())
    if resolved.startswith("\\\\?\\"):
        return resolved
    if len(resolved) >= 240:
        return "\\\\?\\" + resolved
    return resolved


def read_page_text(reader: PdfReader, page_indices: Iterable[int]) -> list[tuple[int, str]]:
    items: list[tuple[int, str]] = []
    for page_idx in page_indices:
        if page_idx < 0 or page_idx >= len(reader.pages):
            continue
        try:
            text = reader.pages[page_idx].extract_text() or ""
        except Exception:
            text = ""
        text = clean_text(text)
        if text:
            items.append((page_idx + 1, text))
    return items


def extract_abstract(text: str) -> str:
    if not text:
        return ""
    match = ABSTRACT_START_RE.search(text)
    if not match:
        return " ".join(split_sentences(text)[:4])
    remaining = text[match.end() :]
    end_match = ABSTRACT_END_RE.search(remaining)
    abstract = remaining[: end_match.start()] if end_match else remaining[:2600]
    return " ".join(split_sentences(abstract)[:6])


def extract_conclusion(text: str) -> str:
    if not text:
        return ""
    match = SECTION_RE.search(text)
    if match:
        text = text[match.end() : match.end() + 3600]
    return " ".join(split_sentences(text)[:8])


def infer_relevance(title: str, abstract: str) -> str:
    low = f"{title} {abstract}".lower()
    score = sum(1 for term in HIGH_RELEVANCE_TERMS if term in low)
    if score >= 4:
        return "high"
    if score >= 2:
        return "medium"
    return "low"


def infer_research_task(title: str, abstract: str) -> str:
    title_low = title.lower()
    if any(key in title_low for key in ["benchmark", "dataset", "survey", "roadmap"]):
        return "综述/数据/基准"
    if any(key in title_low for key in ["navigation", "vln", "nav"]):
        return "规划/导航"
    if any(key in title_low for key in ["ground", "referring", "grasp", "segmentation"]):
        return "指代理解/抓取"
    if any(key in title_low for key in ["clarif", "question", "dialog"]):
        return "交互澄清"
    low = f"{title} {abstract}".lower()
    if any(key in low for key in ["disambigu", "ambigu", "vague", "fuzzy"]):
        return "模糊指令消歧"
    if any(key in low for key in ["planning", "planner", "navigation", "nav"]):
        return "规划/导航"
    if any(key in low for key in ["ground", "referring", "grasp", "segmentation"]):
        return "指代理解/抓取"
    if any(key in low for key in ["clarif", "question", "dialog"]):
        return "交互澄清"
    return "具身理解/控制"


def infer_theme(record_text: str) -> str:
    if "微调/参数高效适配" in record_text:
        return "微调"
    if "对话式澄清" in record_text:
        return "澄清交互"
    if "不确定度驱动" in record_text:
        return "不确定度"
    if "场景图/空间推理" in record_text:
        return "空间推理"
    if "多模态融合" in record_text:
        return "多模态"
    if "数据集/基准" in record_text:
        return "数据集/基准"
    return "通用具身方法"


def summarize_dataset(text: str) -> str:
    picks = pick_sentences(text, DATASET_TERMS, limit=2)
    if picks:
        return trim_text(" ".join(picks), 240)
    return "摘要/前几页中未清楚写出数据集或验证环境。"


def summarize_metrics(text: str) -> str:
    low = text.lower()
    metrics = [term.upper() if term in {"f1", "iou", "miou"} else term for term in METRIC_TERMS if term in low]
    if metrics:
        uniq = []
        for metric in metrics:
            if metric not in uniq:
                uniq.append(metric)
        return ", ".join(uniq[:6])
    metric_sentences = pick_sentences(text, METRIC_TERMS, limit=2)
    if metric_sentences:
        return trim_text(" ".join(metric_sentences), 200)
    return "摘要/前几页中未明确写出核心指标。"


def summarize_limitation(abstract: str, conclusion: str, tail_text: str) -> tuple[str, str]:
    explicit = pick_sentences(
        " ".join([conclusion, tail_text]),
        [
            "limitation",
            "limitations",
            "future work",
            "however",
            "remain",
            "still",
            "challenge",
            "small",
            "simulation",
            "real-world",
        ],
        limit=3,
    )
    if explicit:
        evidence = " ".join(explicit[:2])
        return trim_text(evidence, 260), trim_text(evidence, 320)

    fallback = pick_sentences(
        abstract,
        ["simulation", "dataset", "tabletop", "collected", "single", "few", "real-world"],
        limit=2,
    )
    if fallback:
        evidence = " ".join(fallback[:2])
        return "验证环境看起来偏有限，泛化到复杂工业场景仍需谨慎。", trim_text(evidence, 320)
    return "未在自动抽取到的摘要/结论中看到明确局限，建议后续精读实验部分。", ""


def summarize_method(abstract: str, intro_text: str) -> tuple[str, str]:
    picks = pick_sentences(
        " ".join([abstract, intro_text]),
        [
            "we propose",
            "we present",
            "we introduce",
            "our framework",
            "our method",
            "to tackle this",
            "this paper",
            "we develop",
            "we formulate",
        ],
        limit=2,
    )
    if not picks:
        picks = split_sentences(abstract)[:2] or split_sentences(intro_text)[:2]
    evidence = " ".join(picks[:2])
    return trim_text(evidence, 260), trim_text(evidence, 320)


def infer_inspiration(
    task_scene: str,
    ambiguity_type: str,
    disambiguation_method: str,
    model_architecture: str,
    limitation_summary: str,
) -> str:
    low = f"{task_scene} {ambiguity_type} {disambiguation_method} {model_architecture} {limitation_summary}".lower()
    if "微调/参数高效适配" in disambiguation_method:
        return "可优先参考其参数高效适配思路，把任务缩成工业场景里的小规模指令消歧。"
    if "对话式澄清" in disambiguation_method:
        return "可把“是否需要追问”作为一级分类，再用 LoRA 只微调澄清生成或澄清后的目标排序。"
    if "不确定度驱动" in disambiguation_method:
        return "可把不确定度作为触发条件，只在高风险样本上触发澄清，节省推理和交互成本。"
    if "场景图/空间推理" in disambiguation_method:
        return "可引入空间关系编码或场景图，再用 LoRA 学习工业机械臂里的空间歧义。"
    if "桌面抓取/搬运" in task_scene or "工业机械臂/装配" in task_scene:
        return "可把任务先定成“候选物体排序/目标槽位预测”，避免第一次论文就做全流程控制。"
    if "综述/数据/基准" in low:
        return "可借其类别定义和评测协议，快速搭一个小型工业模糊指令基准。"
    return "建议把任务收缩为“模糊指令检测 + 候选目标排序”，这是最适合一个月内完成的 LoRA 切入点。"


def process_pdf(path: Path, index: int) -> PaperRecord:
    title = derive_title_from_filename(path.name)
    short_title = derive_short_title(title)
    year = year_from_filename(path.name)
    try:
        reader = PdfReader(windows_safe_path(path))
        front_pages = read_page_text(reader, range(min(5, len(reader.pages))))
        tail_start = max(0, len(reader.pages) - 3)
        tail_pages = read_page_text(reader, range(tail_start, len(reader.pages)))

        front_text = " ".join(text for _, text in front_pages)
        tail_text = " ".join(text for _, text in tail_pages)
        abstract = extract_abstract(front_text)
        intro_text = " ".join(split_sentences(front_text)[:6])
        conclusion = extract_conclusion(tail_text)

        title_low = title.lower()
        task_scene = score_rule(f"{title} {abstract}", SCENE_RULES, "通用具身场景", max_labels=1)
        ambiguity_type = score_rule(f"{title} {abstract}", AMBIGUITY_RULES, "未明确/非歧义主任务", max_labels=2)
        disambiguation_rules = DISAMBIGUATION_METHOD_RULES
        architecture_rules = ARCHITECTURE_RULES
        if not any(key in title_low for key in ["benchmark", "dataset", "survey", "roadmap"]):
            disambiguation_rules = [rule for rule in DISAMBIGUATION_METHOD_RULES if rule[0] != "数据集/基准"]
            architecture_rules = [rule for rule in ARCHITECTURE_RULES if rule[0] != "综述/数据资源"]
        disambiguation_method = score_rule(f"{title} {abstract} {conclusion}", disambiguation_rules, "未从摘要中明确识别", max_labels=2)
        modality_input = score_rule(f"{title} {abstract}", MODALITY_RULES, "文本/场景输入未明确", max_labels=2)
        model_architecture = score_rule(f"{title} {abstract}", architecture_rules, "模型架构未明确", max_labels=1)
        dataset_summary = summarize_dataset(" ".join([abstract, front_text]))
        core_metrics = summarize_metrics(" ".join([abstract, front_text]))
        limitation_summary, limitation_evidence = summarize_limitation(abstract, conclusion, tail_text)
        method_summary, method_evidence = summarize_method(abstract, intro_text)
        relevance = infer_relevance(title, abstract)
        research_task = infer_research_task(title, abstract)
        theme = infer_theme(disambiguation_method)
        inspiration = infer_inspiration(task_scene, ambiguity_type, disambiguation_method, model_architecture, limitation_summary)
        pages = []
        if front_pages:
            pages.append(f"front@{front_pages[0][0]}-{front_pages[-1][0]}")
        if tail_pages:
            pages.append(f"tail@{tail_pages[0][0]}-{tail_pages[-1][0]}")
        parse_status = "ok" if abstract else "partial"
        notes = "" if abstract else "摘要未稳定抽取，字段可靠性下降。"

        return PaperRecord(
            index=index,
            batch_id=math.ceil(index / 20),
            filename=path.name,
            full_title=title,
            short_title=short_title,
            year=year,
            task_scene=task_scene,
            ambiguity_type=ambiguity_type,
            disambiguation_method=disambiguation_method,
            modality_input=modality_input,
            model_architecture=model_architecture,
            dataset_summary=dataset_summary,
            core_metrics=core_metrics,
            limitation_summary=limitation_summary,
            inspiration=inspiration,
            theme=theme,
            relevance=relevance,
            research_task=research_task,
            abstract_excerpt=trim_text(abstract, 560),
            method_evidence=method_evidence,
            limitation_evidence=limitation_evidence,
            evidence_pages=",".join(pages),
            parse_status=parse_status,
            notes=notes,
        )
    except Exception as exc:
        return PaperRecord(
            index=index,
            batch_id=math.ceil(index / 20),
            filename=path.name,
            full_title=title,
            short_title=short_title,
            year=year,
            task_scene="解析失败",
            ambiguity_type="解析失败",
            disambiguation_method="解析失败",
            modality_input="解析失败",
            model_architecture="解析失败",
            dataset_summary="PDF 文本层异常，未能自动抽取。",
            core_metrics="PDF 文本层异常，未能自动抽取。",
            limitation_summary="PDF 文本层异常，未能自动抽取。",
            inspiration="先排查该 PDF 是否可被正常解析。",
            theme="未解析",
            relevance="unknown",
            research_task="未解析",
            abstract_excerpt="",
            method_evidence="",
            limitation_evidence="",
            evidence_pages="",
            parse_status="failed",
            notes=f"{type(exc).__name__}: {exc}",
        )


def normalize_title_row(sheet) -> None:
    fill = PatternFill(fill_type="solid", fgColor="D9EAF7")
    for cell in sheet[1]:
        cell.font = Font(bold=True)
        cell.fill = fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = sheet.dimensions


def auto_width(sheet) -> None:
    for idx, column in enumerate(sheet.columns, start=1):
        values = [str(cell.value) if cell.value is not None else "" for cell in column]
        width = min(max((len(v) for v in values), default=10) + 2, 60)
        width = max(width, 12)
        sheet.column_dimensions[get_column_letter(idx)].width = width
    for row in sheet.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)


def safe_value(value):
    if isinstance(value, str):
        return ILLEGAL_EXCEL_RE.sub("", value)
    return value


def add_chart(sheet, title: str, min_col: int, max_col: int, max_row: int, anchor: str) -> None:
    chart = LineChart()
    chart.title = title
    chart.y_axis.title = "论文数"
    chart.x_axis.title = "年份"
    data = Reference(sheet, min_col=min_col, max_col=max_col, min_row=1, max_row=max_row)
    cats = Reference(sheet, min_col=1, min_row=2, max_row=max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 9
    chart.width = 16
    sheet.add_chart(chart, anchor)


def build_main_sheet(workbook: Workbook, records: list[PaperRecord]) -> None:
    ws = workbook.active
    ws.title = "paper_matrix"
    headers = [
        "序号",
        "批次",
        "论文名简写",
        "完整标题",
        "年份",
        "任务场景",
        "歧义类型",
        "消歧方法",
        "模态输入",
        "模型架构",
        "数据集/验证环境",
        "核心指标",
        "局限性",
        "给我的启发",
        "研究任务",
        "主题",
        "相关性",
        "摘要摘录",
        "方法证据",
        "问题证据",
        "证据页码",
        "解析状态",
        "备注",
        "文件名",
    ]
    ws.append(headers)
    for record in records:
        ws.append(
            [
                safe_value(record.index),
                safe_value(record.batch_id),
                safe_value(record.short_title),
                safe_value(record.full_title),
                safe_value(record.year),
                safe_value(record.task_scene),
                safe_value(record.ambiguity_type),
                safe_value(record.disambiguation_method),
                safe_value(record.modality_input),
                safe_value(record.model_architecture),
                safe_value(record.dataset_summary),
                safe_value(record.core_metrics),
                safe_value(record.limitation_summary),
                safe_value(record.inspiration),
                safe_value(record.research_task),
                safe_value(record.theme),
                safe_value(record.relevance),
                safe_value(record.abstract_excerpt),
                safe_value(record.method_evidence),
                safe_value(record.limitation_evidence),
                safe_value(record.evidence_pages),
                safe_value(record.parse_status),
                safe_value(record.notes),
                safe_value(record.filename),
            ]
        )
    normalize_title_row(ws)
    auto_width(ws)


def build_batch_sheet(workbook: Workbook, records: list[PaperRecord]) -> None:
    ws = workbook.create_sheet("batch_review")
    ws.append(["批次", "论文数", "主方法分布", "主要场景", "相关性分布", "解析状态", "复核说明"])
    groups: dict[int, list[PaperRecord]] = defaultdict(list)
    for record in records:
        groups[record.batch_id].append(record)
    for batch_id in sorted(groups):
        items = groups[batch_id]
        method_counter = Counter(item.disambiguation_method for item in items)
        scene_counter = Counter(item.task_scene for item in items)
        relevance_counter = Counter(item.relevance for item in items)
        parse_counter = Counter(item.parse_status for item in items)
        ws.append(
            [
                batch_id,
                len(items),
                ", ".join(f"{k}:{v}" for k, v in method_counter.most_common(4)),
                ", ".join(f"{k}:{v}" for k, v in scene_counter.most_common(4)),
                ", ".join(f"{k}:{v}" for k, v in relevance_counter.items()),
                ", ".join(f"{k}:{v}" for k, v in parse_counter.items()),
                "每 20 篇做一次字段一致性检查；优先核对标题、方法、局限是否与摘要/结论证据一致。",
            ]
        )
    normalize_title_row(ws)
    auto_width(ws)


def build_method_tree_sheet(workbook: Workbook, records: list[PaperRecord]) -> None:
    ws = workbook.create_sheet("method_tree")
    ws.append(["一级方法类", "论文数", "代表论文简写"])
    groups: dict[str, list[PaperRecord]] = defaultdict(list)
    for record in records:
        groups[record.disambiguation_method].append(record)
    for method, items in sorted(groups.items(), key=lambda item: (-len(item[1]), item[0])):
        reps = ", ".join(item.short_title for item in items[:5])
        ws.append([method, len(items), reps])
    normalize_title_row(ws)
    auto_width(ws)

    chart = BarChart()
    chart.title = "方法分支论文数量"
    chart.y_axis.title = "论文数"
    data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
    cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.height = 10
    chart.width = 18
    ws.add_chart(chart, "E2")


def build_year_trend_sheet(workbook: Workbook, records: list[PaperRecord]) -> None:
    ws = workbook.create_sheet("year_trend")
    years = sorted({record.year for record in records if record.year})
    methods = sorted({record.disambiguation_method for record in records})
    ws.append(["年份", *methods])
    for year in years:
        row = [year]
        for method in methods:
            count = sum(1 for record in records if record.year == year and record.disambiguation_method == method)
            row.append(count)
        ws.append(row)
    normalize_title_row(ws)
    auto_width(ws)
    if ws.max_row > 2 and ws.max_column > 1:
        add_chart(ws, "按年份的方法趋势", 2, ws.max_column, ws.max_row, "J2")


def build_gap_sheet(workbook: Workbook, records: list[PaperRecord]) -> None:
    ws = workbook.create_sheet("gap_analysis")
    ws.append(["任务场景", "消歧方法", "现有论文数", "判断", "说明"])
    scene_method_counter = Counter((record.task_scene, record.disambiguation_method) for record in records)
    for scene, method in GAP_CANDIDATES:
        count = scene_method_counter.get((scene, method), 0)
        if count >= 6:
            label = "饱和区"
            note = "本地论文池里同类组合较多，除非评测或数据明显更强，否则不建议首选。"
        elif count >= 2:
            label = "薄弱区"
            note = "已有工作但不算密集，可以考虑做更强基线或更贴近工业的版本。"
        else:
            label = "机会区"
            note = "本地论文池里很少见，适合作为你后续收敛选题的优先候选。"
        ws.append([scene, method, count, label, note])
    normalize_title_row(ws)
    auto_width(ws)


def build_topic_sheet(workbook: Workbook, records: list[PaperRecord]) -> None:
    ws = workbook.create_sheet("candidate_topics")
    ws.append(["候选题目", "研究假设", "新颖性", "数据可得性", "算力需求", "一个月可完成性", "理由"])

    candidates = [
        (
            "基于 LoRA 的工业机械臂模糊指令目标排序方法",
            "LoRA 能让通用 VLM 更好适配工业场景中的目标指代歧义，从而提升候选目标排序准确率。",
            "中高",
            "中",
            "中",
            "高",
            "任务边界小，最适合没有实物设备、需要一个月内完成的第一次论文。",
        ),
        (
            "面向工业机械臂的模糊指令检测与澄清触发策略",
            "在指令执行前先做模糊检测和不确定度估计，可以减少错误执行并降低不必要的交互次数。",
            "中",
            "中高",
            "低",
            "高",
            "比直接做完整消歧更稳，适合先做分类任务再扩展到澄清模块。",
        ),
        (
            "融合空间关系编码与 LoRA 的工业场景指令消歧",
            "显式空间关系特征可弥补通用模型对工业桌面布局理解不足的问题。",
            "中高",
            "中",
            "中",
            "中",
            "更偏研究性，适合你在第一题基础上做强化版本。",
        ),
        (
            "小样本工业模糊指令基准构建与参数高效适配评测",
            "即使缺少真实机械臂，构造小型工业模糊指令数据集并比较 LoRA/Prompt/冻结基线也能形成论文。",
            "中高",
            "高",
            "低中",
            "高",
            "最现实，数据和评测都可控，对第一次发 2/3 区更友好。",
        ),
    ]
    for row in candidates:
        ws.append(list(row))
    normalize_title_row(ws)
    auto_width(ws)


def build_summary_sheet(workbook: Workbook, records: list[PaperRecord]) -> None:
    ws = workbook.create_sheet("summary")
    ws.append(["指标", "数值"])
    relevance_counter = Counter(record.relevance for record in records)
    theme_counter = Counter(record.theme for record in records)
    parse_counter = Counter(record.parse_status for record in records)
    ws.append(["论文总数", len(records)])
    ws.append(["高相关论文", relevance_counter.get("high", 0)])
    ws.append(["中相关论文", relevance_counter.get("medium", 0)])
    ws.append(["低相关论文", relevance_counter.get("low", 0)])
    ws.append(["解析成功", parse_counter.get("ok", 0)])
    ws.append(["部分解析", parse_counter.get("partial", 0)])
    ws.append(["解析失败", parse_counter.get("failed", 0)])
    ws.append(["主题分布", ", ".join(f"{k}:{v}" for k, v in theme_counter.most_common(8))])
    ws.append(["说明", "字段优先绑定到摘要和结论证据；“给我的启发”和候选题目是面向你当前约束的研究建议。"])
    normalize_title_row(ws)
    auto_width(ws)


def build_workbook(records: list[PaperRecord], output_path: Path) -> None:
    workbook = Workbook()
    build_main_sheet(workbook, records)
    build_batch_sheet(workbook, records)
    build_method_tree_sheet(workbook, records)
    build_year_trend_sheet(workbook, records)
    build_gap_sheet(workbook, records)
    build_topic_sheet(workbook, records)
    build_summary_sheet(workbook, records)
    workbook.save(output_path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Build a literature review workbook from local PDF files.")
    parser.add_argument("--root", type=Path, default=Path.cwd(), help="Project root.")
    parser.add_argument("--start", type=int, default=1, help="1-based start index for batch processing.")
    parser.add_argument("--limit", type=int, default=0, help="Number of PDFs to process. 0 means all remaining PDFs.")
    parser.add_argument(
        "--output",
        type=Path,
        default=Path.cwd().parent / f"具身模糊指令文献矩阵_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        help="Output xlsx path.",
    )
    args = parser.parse_args()

    pdf_dir = find_pdf_folder(args.root)
    pdf_files = sorted([path for path in pdf_dir.iterdir() if path.suffix.lower() == ".pdf"], key=lambda path: path.name.lower())
    start_idx = max(args.start, 1)
    start_offset = start_idx - 1
    end_offset = None if args.limit <= 0 else start_offset + args.limit
    selected = pdf_files[start_offset:end_offset]
    records = [process_pdf(path, idx) for idx, path in enumerate(selected, start=start_idx)]
    build_workbook(records, args.output)
    print(f"Processed {len(records)} PDFs")
    print(f"Workbook saved to: {args.output}")


if __name__ == "__main__":
    main()
